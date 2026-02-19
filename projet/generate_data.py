"""
Génère les 4 fichiers Excel du projet Service technique communal d'Yverdon-les-Bains.
Les données contiennent des tendances exploitables par chaque brief.
"""

import random
import datetime
from openpyxl import Workbook

random.seed(42)

# --- Géographie d'Yverdon-les-Bains ---
QUARTIERS = {
    "Gare": {
        "lieux": [
            ("Avenue de la Gare", 46.7808, 6.6418),
            ("Place de la Gare", 46.7813, 6.6413),
            ("Rue de la Plaine", 46.7800, 6.6430),
            ("Passage de l'Hôtel de Ville", 46.7805, 6.6405),
        ]
    },
    "Centre": {
        "lieux": [
            ("Place Pestalozzi", 46.7784, 6.6398),
            ("Rue du Lac", 46.7790, 6.6375),
            ("Rue du Milieu", 46.7788, 6.6402),
            ("Rue du Casino", 46.7795, 6.6388),
            ("Rue de la Maison Rouge", 46.7782, 6.6410),
        ]
    },
    "Rives": {
        "lieux": [
            ("Parc des Rives", 46.7840, 6.6375),
            ("Quai de Nogent", 46.7845, 6.6360),
            ("Plage d'Yverdon", 46.7852, 6.6345),
            ("Avenue des Bains", 46.7835, 6.6355),
        ]
    },
    "Sud": {
        "lieux": [
            ("Y-Parc", 46.7650, 6.6280),
            ("Rue des Pêcheurs", 46.7680, 6.6350),
            ("Centre sportif", 46.7675, 6.6380),
            ("Route de Lausanne", 46.7660, 6.6410),
        ]
    },
    "Est": {
        "lieux": [
            ("HEIG-VD", 46.7817, 6.6590),
            ("Chemin de Maillefer", 46.7810, 6.6560),
            ("Rue Haldimand", 46.7795, 6.6480),
            ("Avenue des Sports", 46.7780, 6.6520),
        ]
    },
}


def rand_offset():
    """Petit décalage GPS réaliste."""
    return random.uniform(-0.0008, 0.0008)


def format_date_messy(dt, style=None):
    """Formate une date de manière variable comme un humain."""
    if style is None:
        style = random.choice(["ch", "iso", "texte", "annee"])
    if style == "ch":
        return dt.strftime("%d.%m.%Y")
    elif style == "iso":
        return dt.strftime("%Y-%m-%d")
    elif style == "texte":
        mois = [
            "", "janvier", "février", "mars", "avril", "mai", "juin",
            "juillet", "août", "septembre", "octobre", "novembre", "décembre",
        ]
        return f"{mois[dt.month]} {dt.year}"
    else:
        return str(dt.year)


def messy_type(base, variants):
    """Retourne une variante réaliste d'un type."""
    return random.choice(variants)


# ========================================================
# 1. INVENTAIRE MOBILIER (~120 items)
# ========================================================
def generate_inventaire():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaire"
    ws.append(["id", "type", "materiau", "lieu", "latitude", "longitude",
               "date_installation", "etat", "remarques"])

    items = []
    item_id = 1

    # Styles d'ID qui évoluent au fil du temps (Jean-Marc a changé sa méthode)
    def make_id(prefix, num, year):
        if year < 2020:
            return f"{prefix}-{num:03d}"
        elif year < 2022:
            return f"{prefix}_{num}"
        else:
            return str(item_id + 1000)

    # --- Bancs (40) ---
    for i in range(40):
        quartier = random.choice(list(QUARTIERS.keys()))
        lieu_name, lat, lng = random.choice(QUARTIERS[quartier]["lieux"])
        year = random.randint(2015, 2023)
        dt = datetime.date(year, random.randint(1, 12), random.randint(1, 28))
        materiau = "bois" if i < 15 else random.choice(["métal", "metal", "Métal"])

        # Trend Brief D: bancs bois = plus usés
        if materiau == "bois" and year < 2020:
            etat = random.choice(["usé", "à remplacer", "usé", "bon"])
        elif materiau == "bois":
            etat = random.choice(["bon", "usé", "bon"])
        else:
            etat = random.choice(["bon", "bon", "bon", "usé"])

        type_val = messy_type("banc", ["banc", "Banc", "banc", "banc public", "Banc"])
        id_val = make_id("B", i + 1, year)

        remarque = ""
        if random.random() < 0.15:
            remarque = random.choice([
                "offert par le Lions Club",
                "à côté de l'arrêt de bus",
                "face au lac",
                "sous un arbre",
                "plaque commémorative",
                "dossier gravé",
            ])

        lat_val = lat + rand_offset()
        lng_val = lng + rand_offset()
        # Quelques GPS arrondis (copié vite fait depuis Google Maps)
        if random.random() < 0.2:
            lat_val = round(lat_val, 2)
            lng_val = round(lng_val, 2)
        # Quelques GPS manquants
        if random.random() < 0.08:
            lat_val = None
            lng_val = None

        items.append({
            "id": id_val, "type": type_val, "materiau": materiau,
            "lieu": lieu_name, "lat": lat_val, "lng": lng_val,
            "date": format_date_messy(dt, "ch" if year < 2020 else None),
            "etat": etat, "remarque": remarque, "quartier": quartier,
            "real_type": "banc", "install_year": year,
        })
        item_id += 1

    # --- Lampadaires (35) ---
    for i in range(35):
        # Trend Brief C: vieux lampadaires concentrés à Gare
        if i < 12:
            quartier = "Gare"
            year = random.randint(2012, 2016)
        elif i < 20:
            quartier = random.choice(["Centre", "Rives"])
            year = random.randint(2016, 2020)
        else:
            quartier = random.choice(["Sud", "Est", "Rives"])
            year = random.randint(2019, 2023)

        lieu_name, lat, lng = random.choice(QUARTIERS[quartier]["lieux"])
        dt = datetime.date(year, random.randint(1, 12), random.randint(1, 28))

        if year < 2018:
            type_val = messy_type("lampadaire", ["lampadaire", "Lampadaire", "lampadaire sodium"])
            materiau = "sodium"
            etat = random.choice(["usé", "usé", "bon", "à remplacer"])
        else:
            type_val = messy_type("lampadaire", ["Lampadaire LED", "lampadaire LED", "lampadaire led"])
            materiau = "LED"
            etat = random.choice(["bon", "bon", "bon", "usé"])

        id_val = make_id("L", i + 1, year)
        lat_val = lat + rand_offset() if random.random() > 0.05 else None
        lng_val = lng + rand_offset() if lat_val else None

        remarque = ""
        if random.random() < 0.1:
            remarque = random.choice(["éclairage passage piéton", "mât 6m", "mât 8m", "double tête"])

        items.append({
            "id": id_val, "type": type_val, "materiau": materiau,
            "lieu": lieu_name, "lat": lat_val, "lng": lng_val,
            "date": format_date_messy(dt),
            "etat": etat, "remarque": remarque, "quartier": quartier,
            "real_type": "lampadaire", "install_year": year,
        })
        item_id += 1

    # --- Poubelles (25) ---
    for i in range(25):
        quartier = random.choice(list(QUARTIERS.keys()))
        lieu_name, lat, lng = random.choice(QUARTIERS[quartier]["lieux"])
        year = random.randint(2016, 2023)
        dt = datetime.date(year, random.randint(1, 12), random.randint(1, 28))
        type_val = messy_type("poubelle", ["poubelle", "Poubelle", "corbeille", "poubelle tri"])
        id_val = make_id("P", i + 1, year)

        items.append({
            "id": id_val, "type": type_val, "materiau": random.choice(["métal", "metal"]),
            "lieu": lieu_name, "lat": lat + rand_offset(), "lng": lng + rand_offset(),
            "date": format_date_messy(dt),
            "etat": random.choice(["bon", "bon", "usé"]),
            "remarque": "", "quartier": quartier,
            "real_type": "poubelle", "install_year": year,
        })
        item_id += 1

    # --- Fontaines (10) --- Trend Brief A: coûteuses
    for i in range(10):
        quartier = random.choice(["Centre", "Rives", "Gare"])
        lieu_name, lat, lng = random.choice(QUARTIERS[quartier]["lieux"])
        year = random.randint(2010, 2018)
        dt = datetime.date(year, random.randint(1, 12), random.randint(1, 28))
        type_val = messy_type("fontaine", ["fontaine", "Fontaine", "fontaine publique"])
        id_val = make_id("F", i + 1, year)

        items.append({
            "id": id_val, "type": type_val, "materiau": random.choice(["pierre", "Pierre", "béton"]),
            "lieu": lieu_name, "lat": lat + rand_offset(), "lng": lng + rand_offset(),
            "date": format_date_messy(dt, "annee" if year < 2015 else "ch"),
            "etat": random.choice(["bon", "usé", "bon"]),
            "remarque": random.choice(["", "", "eau potable", "historique", ""]),
            "quartier": quartier,
            "real_type": "fontaine", "install_year": year,
        })
        item_id += 1

    # --- Bornes recharge (5) ---
    for i in range(5):
        quartier = random.choice(["Gare", "Sud", "Centre"])
        lieu_name, lat, lng = random.choice(QUARTIERS[quartier]["lieux"])
        year = random.randint(2021, 2023)
        dt = datetime.date(year, random.randint(1, 12), random.randint(1, 28))
        id_val = make_id("EV", i + 1, year)

        items.append({
            "id": id_val, "type": random.choice(["borne recharge EV", "Borne recharge", "borne EV"]),
            "materiau": "",
            "lieu": lieu_name, "lat": lat + rand_offset(), "lng": lng + rand_offset(),
            "date": format_date_messy(dt, "iso"),
            "etat": "bon",
            "remarque": random.choice(["22kW", "50kW rapide", ""]),
            "quartier": quartier,
            "real_type": "borne", "install_year": year,
        })
        item_id += 1

    # --- Panneaux affichage (5) ---
    for i in range(5):
        quartier = random.choice(["Centre", "Gare"])
        lieu_name, lat, lng = random.choice(QUARTIERS[quartier]["lieux"])
        year = random.randint(2018, 2022)
        dt = datetime.date(year, random.randint(1, 12), random.randint(1, 28))
        id_val = make_id("PA", i + 1, year)

        items.append({
            "id": id_val, "type": random.choice(["panneau affichage", "Panneau", "panneau info"]),
            "materiau": random.choice(["métal", "bois"]),
            "lieu": lieu_name, "lat": lat + rand_offset(), "lng": lng + rand_offset(),
            "date": format_date_messy(dt),
            "etat": random.choice(["bon", "bon", "usé"]),
            "remarque": "", "quartier": quartier,
            "real_type": "panneau", "install_year": year,
        })
        item_id += 1

    # Ajouter un doublon (même objet entré 2x avec légère variation)
    dup = items[5].copy()
    dup["id"] = items[5]["id"].replace("-", "_")
    dup["lieu"] = items[5]["lieu"].lower()
    items.append(dup)

    # Écriture
    for it in items:
        ws.append([
            it["id"], it["type"], it["materiau"], it["lieu"],
            it["lat"], it["lng"], it["date"], it["etat"], it["remarque"],
        ])

    wb.save("projet/data/inventaire_mobilier.xlsx")
    print(f"inventaire_mobilier.xlsx : {len(items)} lignes")
    return items


# ========================================================
# 2. SIGNALEMENTS (~200)
# ========================================================
def generate_signalements(inventaire):
    wb = Workbook()
    ws = wb.active
    ws.title = "Signalements"
    ws.append(["date", "signale_par", "objet", "description", "urgence", "statut"])

    signalements = []

    # Créer des descriptions réalistes par type
    descriptions = {
        "banc": [
            "Le dossier est fendu",
            "Lattes cassées",
            "Tags / graffitis",
            "Vis qui dépassent",
            "Bois pourri côté gauche",
            "Un pied est tordu",
        ],
        "lampadaire": [
            "Ne s'allume plus",
            "Ampoule grillée",
            "Clignote depuis 1 semaine",
            "Éclaire mal",
            "Mât penché après la tempête",
            "Vitre cassée",
        ],
        "poubelle": [
            "Déborde régulièrement",
            "Couvercle cassé",
            "Renversée",
            "Tags",
            "Brûlée",
        ],
        "fontaine": [
            "Fuit au sol",
            "Ne coule plus",
            "Eau trouble",
            "Mousse/algues",
            "Vasque fissurée",
            "Gel a abîmé la tuyauterie",
        ],
        "borne": [
            "Écran cassé",
            "Ne charge plus",
            "Câble manquant",
        ],
        "panneau": [
            "Vitrine cassée",
            "Affichage décollé",
            "Penché",
        ],
    }

    signale_par_options = [
        "Mme Dupont", "un habitant", "patrouille JM", "M. Keller",
        "email citoyen", "Mme Rochat", "un passant", "concierge école",
        "M. Pereira", "", "habitant du quartier", "Mme Weber",
    ]

    for i in range(200):
        # Trend Brief B: plus de signalements à la Gare
        if random.random() < 0.35:
            candidates = [it for it in inventaire if it["quartier"] == "Gare"]
        else:
            candidates = inventaire

        item = random.choice(candidates)

        # Trend Brief A: pic saisonnier mars-avril et nov-dec
        month_weights = [5, 5, 15, 15, 8, 6, 4, 4, 6, 8, 12, 12]
        month = random.choices(range(1, 13), weights=month_weights)[0]
        year = random.choices([2022, 2023, 2024, 2025], weights=[10, 25, 35, 30])[0]
        day = random.randint(1, 28)
        dt = datetime.date(year, month, day)

        # Date format evolves over time
        if year <= 2022:
            date_str = format_date_messy(dt, "ch")
        elif year <= 2023:
            date_str = format_date_messy(dt, random.choice(["ch", "iso"]))
        else:
            date_str = format_date_messy(dt, "iso")

        # Description: reference l'objet par texte libre (pas par ID)
        lieu = item["lieu"]
        real_type = item["real_type"]
        desc = random.choice(descriptions.get(real_type, ["Problème signalé"]))

        # Objet en texte libre — parfois avec typo
        objet_templates = [
            f"{item['type']} {lieu}",
            f"le {item['type']} devant {lieu}",
            f"{item['type']} près de {lieu}",
            f"{real_type} {lieu}",
        ]
        objet = random.choice(objet_templates)
        # Quelques typos naturelles
        if random.random() < 0.05:
            objet = objet.replace("Place", "place").replace("Gare", "Garre")

        urgence = ""
        if real_type == "lampadaire" and "plus" in desc:
            urgence = "urgent"
        elif real_type == "fontaine" and "fuit" in desc.lower():
            urgence = "urgent"
        elif random.random() < 0.3:
            urgence = random.choice(["urgent", "normal"])
        # Souvent vide (Jean-Marc ne remplit pas toujours)

        # Trend Brief B: statuts — Gare a plus de "en attente"
        if item["quartier"] == "Gare" and random.random() < 0.4:
            statut = random.choice(["en attente", "en attente", "en cours", ""])
        elif dt > datetime.date(2025, 1, 1) and random.random() < 0.3:
            statut = random.choice(["en attente", "en cours", ""])
        else:
            statut = random.choice(["fait", "fait", "fait", "en attente", ""])

        signalements.append({
            "date": date_str, "date_obj": dt,
            "signale_par": random.choice(signale_par_options),
            "objet": objet, "description": desc,
            "urgence": urgence, "statut": statut,
            "quartier": item["quartier"],
            "real_type": real_type,
            "item_id": item["id"],
        })

    # Ajouter 3 doublons (même signalement à 1 jour d'écart)
    for dup_idx in [10, 45, 120]:
        dup = signalements[dup_idx].copy()
        orig_dt = signalements[dup_idx]["date_obj"]
        new_dt = orig_dt + datetime.timedelta(days=1)
        dup["date"] = format_date_messy(new_dt, "ch")
        dup["signale_par"] = random.choice(signale_par_options)
        signalements.append(dup)

    # Mélanger un peu
    random.shuffle(signalements)

    for s in signalements:
        ws.append([s["date"], s["signale_par"], s["objet"],
                   s["description"], s["urgence"], s["statut"]])

    wb.save("projet/data/signalements.xlsx")
    print(f"signalements.xlsx : {len(signalements)} lignes")
    return signalements


# ========================================================
# 3. INTERVENTIONS (~150)
# ========================================================
def generate_interventions(inventaire, signalements):
    wb = Workbook()
    ws = wb.active
    ws.title = "Interventions"
    ws.append(["date", "objet", "type_intervention", "technicien",
               "duree", "cout_materiel", "remarques"])

    interventions = []
    techniciens = ["Jean-Marc", "JM", "Jean-Marc Bonvin", "Alves Pedro",
                   "P. Alves", "Pedro", "stagiaire", "Koffi Marc"]

    type_interventions = {
        "banc": ["réparation", "Réparation", "remplacement latte",
                 "peinture", "nettoyage tags"],
        "lampadaire": ["remplacement ampoule", "réparation", "Réparation électrique",
                       "remplacement complet", "redressage mât"],
        "poubelle": ["réparation", "remplacement couvercle", "nettoyage", "remplacement"],
        "fontaine": ["réparation fuite", "détartrage", "hivernage",
                     "remise en service", "remplacement pompe"],
        "borne": ["réparation", "mise à jour logiciel"],
        "panneau": ["réparation vitre", "nettoyage"],
    }

    # Coûts réalistes par type — Trend Brief A: fontaines = chères
    cout_ranges = {
        "banc": (20, 120),
        "lampadaire": (30, 250),
        "poubelle": (15, 80),
        "fontaine": (150, 600),
        "borne": (50, 200),
        "panneau": (20, 100),
    }

    # Trend Brief C: les vieux lampadaires Gare ont plus d'interventions
    vieux_lampadaires_gare = [
        it for it in inventaire
        if it["real_type"] == "lampadaire"
        and it["quartier"] == "Gare"
        and it["install_year"] < 2017
    ]

    # Bancs bois anciens — Trend Brief D
    vieux_bancs_bois = [
        it for it in inventaire
        if it["real_type"] == "banc"
        and it["materiau"] == "bois"
        and it["install_year"] < 2020
    ]

    # Générer interventions pour vieux lampadaires (3-5 chacun)
    for lamp in vieux_lampadaires_gare:
        n_interventions = random.randint(3, 5)
        for _ in range(n_interventions):
            year = random.randint(2020, 2025)
            # Trend saisonnier: lampadaires surtout en automne/hiver
            month = random.choices(range(1, 13),
                                   weights=[8, 6, 4, 3, 2, 2, 2, 3, 6, 10, 12, 12])[0]
            dt = datetime.date(year, month, random.randint(1, 28))
            cout_min, cout_max = cout_ranges["lampadaire"]
            cout = random.randint(cout_min, cout_max)

            interventions.append({
                "date": dt, "objet": f"{lamp['type']} {lamp['lieu']}",
                "type_intervention": random.choice(type_interventions["lampadaire"]),
                "technicien": random.choice(techniciens[:4]),
                "duree": random.choice(["1h", "2h", "1h30", "30 min", "une matinée"]),
                "cout": cout, "real_type": "lampadaire",
                "remarques": random.choice(["", "", "pièce commandée", "à surveiller"]),
            })

    # Interventions pour vieux bancs bois (2-3 chacun)
    for banc in vieux_bancs_bois:
        n_interventions = random.randint(2, 4)
        for _ in range(n_interventions):
            year = random.randint(2021, 2025)
            month = random.choices(range(1, 13),
                                   weights=[4, 4, 15, 15, 8, 5, 3, 3, 5, 8, 8, 4])[0]
            dt = datetime.date(year, month, random.randint(1, 28))
            cout_min, cout_max = cout_ranges["banc"]
            cout = random.randint(cout_min, cout_max)

            interventions.append({
                "date": dt, "objet": f"{banc['type']} {banc['lieu']}",
                "type_intervention": random.choice(type_interventions["banc"]),
                "technicien": random.choice(techniciens),
                "duree": random.choice(["1h", "2h", "30 min", "une matinée", "une journée"]),
                "cout": cout, "real_type": "banc",
                "remarques": random.choice(["", "", "bois humide", "latte fendue"]),
            })

    # Interventions fontaines (chères, régulières)
    fontaines = [it for it in inventaire if it["real_type"] == "fontaine"]
    for fontaine in fontaines:
        n_interventions = random.randint(2, 4)
        for _ in range(n_interventions):
            year = random.randint(2021, 2025)
            # Hivernage en novembre, remise en service en mars
            month = random.choices(range(1, 13),
                                   weights=[3, 3, 15, 8, 5, 5, 3, 3, 5, 8, 15, 5])[0]
            dt = datetime.date(year, month, random.randint(1, 28))
            cout_min, cout_max = cout_ranges["fontaine"]
            cout = random.randint(cout_min, cout_max)

            interventions.append({
                "date": dt, "objet": f"{fontaine['type']} {fontaine['lieu']}",
                "type_intervention": random.choice(type_interventions["fontaine"]),
                "technicien": random.choice(techniciens[:4]),
                "duree": random.choice(["2h", "3h", "une journée", "une matinée"]),
                "cout": cout, "real_type": "fontaine",
                "remarques": random.choice([
                    "", "", "plombier externe",
                    "garantie fournisseur", "pièce spéciale commandée",
                ]),
            })

    # Compléter avec des interventions normales pour le reste
    autres = [it for it in inventaire
              if it not in vieux_lampadaires_gare
              and it not in vieux_bancs_bois
              and it not in fontaines]
    target = 150 - len(interventions)
    for _ in range(max(0, target)):
        item = random.choice(autres)
        year = random.randint(2022, 2025)
        month = random.choices(range(1, 13),
                               weights=[5, 5, 12, 12, 8, 6, 4, 4, 6, 8, 10, 8])[0]
        dt = datetime.date(year, month, random.randint(1, 28))
        rt = item["real_type"]
        cout_min, cout_max = cout_ranges.get(rt, (20, 100))
        cout = random.randint(cout_min, cout_max)

        interventions.append({
            "date": dt,
            "objet": f"{item['type']} {item['lieu']}",
            "type_intervention": random.choice(
                type_interventions.get(rt, ["réparation"])),
            "technicien": random.choice(techniciens),
            "duree": random.choice(["1h", "2h", "30 min", "1h30"]),
            "cout": cout, "real_type": rt,
            "remarques": "",
        })

    random.shuffle(interventions)

    # Écriture avec formats de coûts et dates variables
    for interv in interventions:
        dt = interv["date"]
        # Format date qui évolue
        if dt.year <= 2022:
            date_str = dt.strftime("%d.%m.%Y")
        elif dt.year <= 2023:
            date_str = dt.strftime(random.choice(["%d.%m.%Y", "%Y-%m-%d"]))
        else:
            date_str = dt.strftime("%Y-%m-%d")

        # Format coût variable
        cout = interv["cout"]
        if random.random() < 0.1:
            cout_str = f"CHF {cout}.-"
        elif random.random() < 0.1:
            cout_str = f"{cout}.-"
        elif cout == 0 or (interv["remarques"] == "garantie fournisseur"):
            cout_str = random.choice(["0", "gratuit", "garantie"])
        else:
            cout_str = str(cout)

        # Quelques coûts vides
        if random.random() < 0.08:
            cout_str = ""

        ws.append([
            date_str, interv["objet"], interv["type_intervention"],
            interv["technicien"], interv["duree"], cout_str,
            interv["remarques"],
        ])

    wb.save("projet/data/interventions.xlsx")
    print(f"interventions.xlsx : {len(interventions)} lignes")
    return interventions


# ========================================================
# 4. FOURNISSEURS (~15)
# ========================================================
def generate_fournisseurs():
    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseurs"
    ws.append(["entreprise", "contact", "telephone", "email",
               "type_materiel", "remarques"])

    fournisseurs = [
        ("Mobilier Urbain Suisse SA", "M. Steiner", "024 423 12 34",
         "info@mus-sa.ch", "bancs, poubelles", "FERMÉ depuis mars 2025"),
        ("Eclairage Public Romandie", "Keller Thomas", "021 678 90 12",
         "t.keller@epr.ch", "lampadaires, éclairage LED", "bon fournisseur, délai 2 sem."),
        ("Fontaines & Patrimoine Sàrl", "Mme Rochat", "079 345 67 89",
         "", "fontaines, vasques", "artisan, travail soigné mais cher"),
        ("GreenCharge SA", "P. Müller", "+41 21 456 78 90",
         "contact@greencharge.ch", "bornes EV", "contrat maintenance inclus"),
        ("Bächler Metallbau AG", "voir site web", "032 456 78 90",
         "info@baechler.ch", "bancs métal, poubelles", ""),
        ("Signalétique Vaud", "Jean-Paul", "0791234567",
         "jp@signaletique-vaud.ch", "panneaux", ""),
        ("Quincaillerie de la Gare", "", "024 425 00 00",
         "", "visserie, peinture, petites pièces", "pour les urgences"),
        ("Holz Design GmbH", "Hr. Weber", "031 567 89 01",
         "weber@holzdesign.ch", "bancs bois, tables", "germanophone"),
        ("Paysagiste Roth", "Marc Roth", "079 876 54 32",
         "", "plantations autour mobilier", ""),
        ("ABB Suisse", "", "058 585 00 00",
         "voir site web", "éclairage, bornes", "grands projets uniquement"),
        ("Communauté tarifaire vaudoise", "Secrétariat", "021 234 56 78",
         "info@ctv.ch", "abris bus", "pas notre domaine direct"),
        ("Récupération Métal Broye SA", "Alain", "026 789 01 23",
         "", "reprise ancien mobilier", "payent au kilo"),
        ("Ville Propre Sàrl", "Mme Da Silva", "079 111 22 33",
         "contact@villepropre.ch", "nettoyage, tags", "contrat annuel CHF 8000.-"),
        ("Plomberie Noverraz", "Robert", "024 426 33 44",
         "", "fontaines, conduites", "rapide pour les urgences"),
    ]

    for f in fournisseurs:
        ws.append(list(f))

    wb.save("projet/data/fournisseurs_contacts.xlsx")
    print(f"fournisseurs_contacts.xlsx : {len(fournisseurs)} lignes")


# ========================================================
# MAIN
# ========================================================
if __name__ == "__main__":
    print("Génération des données du Service technique Yverdon")
    print("=" * 60)
    inventaire = generate_inventaire()
    signalements = generate_signalements(inventaire)
    generate_interventions(inventaire, signalements)
    generate_fournisseurs()
    print("=" * 60)
    print("Fichiers générés dans projet/data/")
